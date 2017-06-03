from openpyxl import load_workbook
import urllib2
from BeautifulSoup import BeautifulSoup

def get_provider(ip):
    response = urllib2.urlopen('http://ipinfo.io/' + ip)
    html = BeautifulSoup(response.read())
    html.prettify()
    data = html.findAll('td')

    line = BeautifulSoup(str(data[3]))
    print line.text

    return data[3]

try:
    filename = 'D:\\ip.xlsx'
    wb = load_workbook(filename)
    ws = wb.active

    for i in range(2, 3):
        row_ip = ws['A'+str(i)].value
        print(row_ip)
        provider = get_provider(row_ip)
        print provider
        ws['C'+str(i)] = str(provider)

    wb.save(filename)

except Exception as error:
    print error
    wb.save(filename)












